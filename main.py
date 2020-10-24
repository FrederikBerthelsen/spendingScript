import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import re
from datetime import datetime

# Get current Row and Column from file, used to keep track of where to write in .xlsx file. 
# Make sure format is first line = row, second line = col
def getCurrentRowCol(sheet):
    with open(sheet, 'r', newline='') as rowCol:
        try:
            row = int(rowCol.readline().rstrip('\n'))
            col = int(rowCol.readline())
            return row, col
        except:
            raise Exception('currentRowCol file is of wrong type')

# Updates currentRowCol file with given row and col
def updateCurrentRowCol(sheet, row, col):
    with open(sheet, 'r+') as f:
        data = f.read()
        f.seek(0)
        f.write("{0}\n{1}".format(row, col))
        f.truncate()

# Opens CSV file and formats it nicely as lists
def openCSV(filename):
    with open(filename, newline='', encoding='utf-8-sig') as csvfile:
        reader = csv.reader(csvfile, delimiter=';', quotechar='"')
        data = [row for row in reader]
        return data

# Writes titles into spreadsheet, for example, which month and a Date, Text, Amount and total
def writeTitles(xlsxFile, sheet, month):
    wb = load_workbook(xlsxFile)
    ws = wb[sheet]

    row, col = getCurrentRowCol(sheet)

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+4)
    ws.cell(row=row, column=col, value=month)
    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    ws.cell(row=row+1, column=col, value='Date')
    ws.cell(row=row+1, column=col+1, value='Text')
    ws.cell(row=row+1, column=col+2, value='Amount')
    ws.cell(row=row+1, column=col+3, value='Total')
    ws.cell(row=row+1, column=col+4, value='Expenses')

    updateCurrentRowCol(sheet, row+2, col)
    
    wb.save(xlsxFile)
    wb.close()

# Sorts the data from the parameter data. This is expected to be the data output from the CSV file.
# Also makes sure data is only from given month
def sortData(data, month):
    foodExpenses = []
    income = []
    otherExpenses = []
    for purchase in data:
        del purchase[0]
    for purchase in data:
        purchase[2] = float(purchase[2].replace('[A-Za-z]', '').replace('.', '').replace(',', '.'))
        purchase[3] = float(purchase[3].replace('[A-Za-z]', '').replace('.', '').replace(',', '.'))
        if (datetime.strptime(purchase[0], "%d-%m-%Y").strftime("%B")) == month:
            if 'Netto' in purchase[1]:
                foodExpenses.append(purchase)
            elif purchase[2] > 0:
                income.append(purchase)
            else:
                otherExpenses.append(purchase) 
    return foodExpenses, otherExpenses, income

# Writes the data to the spreadsheet
def writeSpecificData(xlsxFile, sheet, month, data):
    writeTitles(xlsxFile, sheet, month)
    row, col = getCurrentRowCol(sheet)
    wb = load_workbook(xlsxFile)
    ws = wb[sheet]
    originalRow = row
    for r in data:
        tempCol = col
        for c in r:
            ws.cell(row=row, column=tempCol, value=c)
            tempCol += 1
        ws.cell(row=row, column=tempCol, value="=SUM({0}:{1})"\
            .format(ws.cell(row=originalRow, column=tempCol-2).coordinate,\
                 ws.cell(row=row, column=tempCol-2).coordinate))
        row += 1
    updateCurrentRowCol(sheet, originalRow  + len(data) + 2, col)
    ws.column_dimensions.bestFit = True
    wb.save(xlsxFile)
    wb.close()

# Calls writeSpecificData for all excel sheets.
def writeAllData(xlsxFile, month):
    data = openCSV("{}.csv".format(month))
    foodExpenses, otherExpenses, income = sortData(data, month)

    writeSpecificData(xlsxFile, 'foodExpenses', month, foodExpenses)
    writeSpecificData(xlsxFile, 'otherExpenses', month, otherExpenses)
    writeSpecificData(xlsxFile, 'income', month, income)

writeAllData('finances.xlsx', input("Month?\n").capitalize())
